"""Сборка книги Excel из датасета: три листа — wide, tidy и параметры запроса."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .query import Dataset, format_period

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
AGGREGATE_FILL = PatternFill("solid", fgColor="EAF1F8")
TITLE_FONT = Font(bold=True, size=12)

FLOW_LABEL = {"I": "Import", "E": "Export"}
INDICATOR_LABEL = {"VAL": "Value", "QTY": "Quantity"}
MIRROR_LABEL = {"D": "Direct", "M": "Mirror"}
OUTPUT_LABEL = {
    "byPartner": "по партнёрам",
    "byProduct": "по продуктам",
    "byCountry": "по странам",
}

# TradeMap помечает отдельные значения числовыми сносками (поле flag).
# Их расшифровка задаётся на стороне сайта и здесь намеренно не дублируется,
# чтобы не выдать неверное толкование: коды выводятся как есть.
FLAG_LEGEND_NOTE = (
    "TradeMap помечает часть значений сносками. Их расшифровка приведена "
    "под таблицей на www.trademap.org для того же среза."
)


def _style_header(sheet, row: int, width: int) -> None:
    for col in range(1, width + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(sheet, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width


def _number_format(indicator: str) -> str:
    return "#,##0" if indicator == "VAL" else "#,##0.###"


def _build_wide(sheet, data: Dataset) -> None:
    spec = data.spec
    lookup = data.lookup()

    base_headers = [
        "Страна-репортёр", "Код", "Партнёр", "Код",
        "Продукт", "Код", "Направление", "Показатель", "Ед. изм.",
    ]
    headers = base_headers + [format_period(p, spec.frequency) for p in data.periods]
    sheet.append(headers)
    _style_header(sheet, 1, len(headers))

    row = 2
    for key in data.keys():
        reporter, partner, product = key
        aggregate = data.is_aggregate(key)
        unit = data.unit_for(key)

        for indicator in spec.indicators:
            values = [
                lookup.get((reporter, partner, product, period, indicator))
                for period in data.periods
            ]
            if not any(values):
                continue

            sheet.cell(row=row, column=1, value=data.labels.country(reporter))
            sheet.cell(row=row, column=2, value=reporter)
            sheet.cell(row=row, column=3, value=data.labels.country(partner))
            sheet.cell(row=row, column=4, value=partner)
            sheet.cell(row=row, column=5, value=data.labels.product(product))
            sheet.cell(row=row, column=6, value=product)
            sheet.cell(row=row, column=7, value=FLOW_LABEL.get(spec.trade_flow, spec.trade_flow))
            sheet.cell(row=row, column=8, value=INDICATOR_LABEL.get(indicator, indicator))
            sheet.cell(
                row=row,
                column=9,
                value=unit if indicator == "QTY" else f"тыс. {spec.currency}",
            )

            for offset, cell in enumerate(values):
                target = sheet.cell(row=row, column=len(base_headers) + 1 + offset)
                if cell is not None:
                    target.value = cell.value
                    target.number_format = _number_format(indicator)

            if aggregate:
                for col in range(1, len(headers) + 1):
                    sheet.cell(row=row, column=col).fill = AGGREGATE_FILL
                    sheet.cell(row=row, column=col).font = Font(bold=True)
            row += 1

    sheet.freeze_panes = sheet.cell(row=2, column=len(base_headers) + 1)
    if row > 2:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    _autosize(sheet, {1: 26, 2: 7, 3: 26, 4: 7, 5: 46, 6: 12, 7: 13, 8: 12, 9: 14})
    for col in range(len(base_headers) + 1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13


def _build_tidy(sheet, data: Dataset) -> None:
    spec = data.spec
    headers = [
        "reporter", "reporterName", "partner", "partnerName",
        "product", "productName", "flow", "frequency", "period", "periodLabel",
        "indicator", "value", "unit", "currency", "flag", "isAggregate",
    ]
    sheet.append(headers)
    _style_header(sheet, 1, len(headers))

    row = 2
    for cell in sorted(
        data.cells,
        key=lambda c: (c.reporter, c.product, c.partner, c.indicator, c.period),
    ):
        sheet.append(
            [
                cell.reporter,
                data.labels.country(cell.reporter),
                cell.partner,
                data.labels.country(cell.partner),
                cell.product,
                data.labels.product(cell.product),
                FLOW_LABEL.get(spec.trade_flow, spec.trade_flow),
                spec.frequency,
                cell.period,
                format_period(cell.period, spec.frequency),
                cell.indicator,
                cell.value,
                cell.unit or ("" if cell.indicator == "QTY" else f"тыс. {spec.currency}"),
                spec.currency if cell.indicator == "VAL" else "",
                cell.flag or "",
                "yes" if cell.is_aggregate else "",
            ]
        )
        sheet.cell(row=row, column=12).number_format = _number_format(cell.indicator)
        row += 1

    sheet.freeze_panes = "A2"
    if row > 2:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    _autosize(
        sheet,
        {1: 10, 2: 26, 3: 10, 4: 26, 5: 12, 6: 46, 7: 10, 8: 11,
         9: 10, 10: 12, 11: 11, 12: 16, 13: 14, 14: 10, 15: 8, 16: 12},
    )


def _build_query_sheet(sheet, data: Dataset) -> None:
    spec = data.spec
    sheet["A1"] = "Параметры выгрузки"
    sheet["A1"].font = TITLE_FONT

    rows = [
        ("Источник", "ITC Trade Map (www.trademap.org)"),
        ("Выгружено (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Частота", spec.frequency),
        ("Режим вывода", f"{spec.output} ({OUTPUT_LABEL.get(spec.output, '')})"),
        ("Направление", FLOW_LABEL.get(spec.trade_flow, spec.trade_flow)),
        ("Показатели", ", ".join(INDICATOR_LABEL.get(i, i) for i in spec.indicators)),
        ("Валюта", spec.currency),
        ("Данные", MIRROR_LABEL.get(spec.direct_mirror, spec.direct_mirror)),
        ("Период с", format_period(spec.period_from, spec.frequency)),
        ("Период по", format_period(spec.period_to, spec.frequency)),
        ("Страны-репортёры", ", ".join(f"{c} {data.labels.country(c)}" for c in spec.reporters) or "—"),
        ("Продукты", ", ".join(f"{p} {data.labels.product(p)}" for p in spec.products) or "—"),
        ("Партнёры", ", ".join(f"{p} {data.labels.country(p)}" for p in spec.partners) or "—"),
        ("Уровень HS", spec.hs_level if spec.output == "byProduct" else "—"),
        ("Значений выгружено", len(data.cells)),
        ("Строк в wide", len(data.keys())),
    ]
    row = 3
    for label, value in rows:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Национальные источники данных").font = TITLE_FONT
    row += 1
    if data.sources:
        for source in data.sources:
            country = data.labels.country(source.get("countryCd") or "")
            text = source.get("labelEn") or ""
            since = source.get("startingDate") or ""
            url = source.get("url") or ""
            sheet.cell(row=row, column=1, value=country)
            sheet.cell(row=row, column=2, value=f"{text}{f' (с {since})' if since else ''}")
            sheet.cell(row=row, column=3, value=url)
            row += 1
    else:
        sheet.cell(row=row, column=2, value="TradeMap не вернул сведений об источниках")
        row += 1

    used_flags = sorted({f for c in data.cells if c.flag for f in set(c.flag)})
    if used_flags:
        row += 1
        sheet.cell(row=row, column=1, value="Сноски в данных").font = TITLE_FONT
        row += 1
        sheet.cell(row=row, column=1, value="Встретились коды")
        sheet.cell(row=row, column=2, value=", ".join(used_flags))
        row += 1
        sheet.cell(row=row, column=2, value=FLAG_LEGEND_NOTE)
        row += 1

    if data.warnings:
        row += 1
        sheet.cell(row=row, column=1, value="Предупреждения").font = TITLE_FONT
        row += 1
        for warning in data.warnings:
            sheet.cell(row=row, column=2, value=warning)
            row += 1

    _autosize(sheet, {1: 28, 2: 90, 3: 40})


def build_workbook(data: Dataset) -> bytes:
    workbook = Workbook()

    wide = workbook.active
    wide.title = "Data (wide)"
    _build_wide(wide, data)

    _build_tidy(workbook.create_sheet("Data (tidy)"), data)
    _build_query_sheet(workbook.create_sheet("Query"), data)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def suggest_filename(data: Dataset) -> str:
    spec = data.spec
    parts = [
        "trademap",
        spec.frequency,
        FLOW_LABEL.get(spec.trade_flow, spec.trade_flow).lower(),
        "-".join(spec.reporters[:3]) or "all",
        "-".join(spec.products[:3]) or "all",
        f"{spec.period_from}_{spec.period_to}",
    ]
    return "_".join(str(p) for p in parts if p) + ".xlsx"
